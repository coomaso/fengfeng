import base64
import json
import os
import random
import time
from datetime import datetime, timezone, timedelta
from typing import Dict, List, Optional, Tuple, Any, Union
from urllib.parse import quote

import requests
from Crypto.Cipher import AES
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


# ==================== 配置常量 ====================
class Config:
    """全局配置"""
    # 网络请求配置
    BASE_URL = "http://106.15.60.27:22222"
    HEADERS = {
        "Accept": "application/json",
        "Accept-Encoding": "gzip, deflate, br",
        "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8,vi;q=0.7",
        "Connection": "keep-alive",
        "Content-Type": "application/json; charset=utf-8",
        "Host": "106.15.60.27:22222",
        "Referer": "http://106.15.60.27:22222/xxgs/",
        "Sec-Ch-Ua": '"Chromium";v="122", "Not(A:Brand";v="24", "Google Chrome";v="122"',
        "Sec-Ch-Ua-Mobile": "?0",
        "Sec-Ch-Ua-Platform": '"Windows"',
        "Sec-Fetch-Dest": "empty",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Site": "same-origin",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.6261.95 Safari/537.36"
    }
    RETRY_COUNT = 3               # 请求重试次数
    PAGE_RETRY_MAX = 2           # 单页最大重试次数
    TIMEOUT = 15                  # 请求超时时间（秒）
    PAGE_SIZE = 10                # 每页记录数

    # AES 加密配置
    AES_KEY = b"6875616E6779696E6875616E6779696E"
    AES_IV = b"sskjKingFree5138"

    # 工作表配置
    COLUMNS = [
        {'id': 'cioName',    'name': '企业名称',   'width': 35,  'merge': True,  'align': 'left'},
        {'id': 'eqtName',    'name': '资质类别',   'width': 20,  'merge': True,  'align': 'center'},
        {'id': 'csf',        'name': '初始分',     'width': 12,  'merge': True,  'align': 'center', 'format': '0'},
        {'id': 'zzmx',       'name': '资质明细',   'width': 50,  'merge': False, 'align': 'left'},
        {'id': 'cxdj',       'name': '诚信等级',   'width': 12,  'merge': False, 'align': 'center'},
        {'id': 'score',      'name': '诚信分值',   'width': 12,  'merge': False, 'align': 'center', 'format': '0.00'},
        {'id': 'jcf',        'name': '基础分',     'width': 12,  'merge': False, 'align': 'center', 'format': '0'},
        {'id': 'zxjf',       'name': '专项加分',   'width': 12,  'merge': False, 'align': 'center', 'format': '0.00'},
        {'id': 'kf',         'name': '扣分',       'width': 12,  'merge': False, 'align': 'center', 'format': '0.00'},
        {'id': 'eqlId',      'name': '资质ID',     'width': 25,  'merge': False, 'align': 'center'},
        {'id': 'orgId',      'name': '组织ID',     'width': 30,  'merge': True,  'align': 'center'},
        {'id': 'cecId',      'name': '信用档案ID', 'width': 30,  'merge': True,  'align': 'center'}
    ]
    SHEET_CONFIGS = [
        {"name": "企业信用数据汇总", "prefix": None, "freeze": 'B2', "merge": True},
        {"name": "建筑工程总承包信用分排序", "prefix": "建筑业企业资质_施工总承包_建筑工程_", "freeze": 'B2', "merge": False, "generate_json": True},
        {"name": "市政公用工程信用分排序", "prefix": "建筑业企业资质_施工总承包_市政公用工程_", "freeze": 'B2', "merge": False, "generate_json": True},
        {"name": "装修装饰工程信用分排序", "prefix": "建筑业企业资质_专业承包_建筑装修装饰工程_", "freeze": 'B2', "merge": False, "generate_json": True},
        {"name": "水利水电工程信用分排序", "prefix": "建筑业企业资质_施工总承包_水利水电工程_", "freeze": 'B2', "merge": False, "generate_json": True},
        {"name": "电力工程信用分排序", "prefix": "建筑业企业资质_施工总承包_电力工程_", "freeze": 'B2', "merge": False, "generate_json": True}
    ]


# ==================== 工具函数 ====================
def safe_request(session: requests.Session, url: str) -> requests.Response:
    """带自动重试的安全请求"""
    for attempt in range(Config.RETRY_COUNT):
        try:
            if attempt > 0:
                time.sleep(random.uniform(0.5, 2.5))
            print(f"正在请求: {url}")
            response = session.get(url, headers=Config.HEADERS, timeout=Config.TIMEOUT)
            response.raise_for_status()
            return response
        except requests.exceptions.Timeout:
            print(f"↺ 请求超时，正在重试 ({attempt+1}/{Config.RETRY_COUNT})...")
        except requests.exceptions.RequestException as e:
            print(f"请求异常: {str(e)}")
            if attempt < Config.RETRY_COUNT - 1:
                print(f"正在进行第 {attempt+2} 次尝试...")
    raise RuntimeError(f"超过最大重试次数 ({Config.RETRY_COUNT})")


def aes_decrypt_base64(encrypted_base64: str) -> str:
    """AES解密函数"""
    if not encrypted_base64:
        raise ValueError("加密数据为空，无法解密")
    try:
        encrypted_bytes = base64.b64decode(encrypted_base64)
        cipher = AES.new(Config.AES_KEY, AES.MODE_CBC, Config.AES_IV)
        decrypted_bytes = cipher.decrypt(encrypted_bytes)
        return decrypted_bytes.rstrip(b'\x00').decode("utf-8")
    except Exception as e:
        print(f"解密失败，原始数据: {encrypted_base64[:50]}...")
        raise RuntimeError(f"解密失败: {str(e)}")


def parse_response_data(encrypted_data: str) -> dict:
    """解密并解析响应数据"""
    if not encrypted_data:
        print("警告: 收到空的加密数据")
        return {"error": "empty data"}
    try:
        decrypted_str = aes_decrypt_base64(encrypted_data)
        print(f"解密后的数据样本: {decrypted_str[:200]}...")
        return json.loads(decrypted_str)
    except json.JSONDecodeError as e:
        print(f"JSON解析错误，数据样本: {encrypted_data[:200]}...")
        return {"error": f"invalid json format: {str(e)}"}
    except Exception as e:
        return {"error": str(e)}


def get_new_code(session: requests.Session) -> Tuple[str, str]:
    """获取新验证码和时间戳"""
    timestamp = str(int(time.time() * 1000))
    code_url = f"{Config.BASE_URL}/ycdc/bakCmisYcOrgan/getCreateCode?codeValue={timestamp}"
    try:
        response = safe_request(session, code_url).json()
        print(f"验证码接口响应: {json.dumps(response, ensure_ascii=False)[:100]}...")
        if response.get("code") != 0:
            raise RuntimeError(f"验证码接口异常: {response}")
        return aes_decrypt_base64(response["data"]), timestamp
    except Exception as e:
        print(f"获取验证码失败，URL: {code_url}")
        raise RuntimeError(f"获取新验证码失败: {str(e)}")


def process_page(session: requests.Session, page: int, code: str, timestamp: str) -> Tuple[List[dict], int]:
    """处理单页数据，包含重试机制"""
    max_retries = 3
    current_code, current_ts = code, timestamp

    for attempt in range(max_retries + 1):
        page_url = (f"{Config.BASE_URL}/ycdc/bakCmisYcOrgan/getCurrentIntegrityPage"
                    f"?pageSize={Config.PAGE_SIZE}&cioName=%E5%85%AC%E5%8F%B8&page={page}"
                    f"&code={quote(current_code)}&codeValue={current_ts}")

        try:
            response = safe_request(session, page_url)
            page_response = response.json()
            status = page_response.get('code', '未知')
            print(f"第 {page} 页 请求#{attempt+1} 响应状态: {status}")

            if "data" not in page_response or not page_response["data"]:
                print(f"空数据响应，准备重试（剩余重试次数: {max_retries - attempt}）")
                if attempt < max_retries:
                    continue
                raise RuntimeError("连续空响应，终止重试")

            page_data = parse_response_data(page_response["data"])
            records = page_data.get("data", [])
            print(f"第 {page} 页解析出 {len(records)} 条记录")
            if not records:
                print(f"警告: 第 {page} 页解析出空记录列表")
            return records, page_data.get("total", 0)
        except Exception as e:
            print(f"第 {page} 页处理失败: {str(e)}")
            raise

    raise RuntimeError("超过最大重试次数")


def fetch_company_detail(session: requests.Session, cec_id: str, company_name: str, max_retries: int = 3) -> dict:
    """获取企业信誉分明细（带重试）"""
    print(f"\n获取企业信誉分明细: {company_name} (cecId={cec_id})")
    detail_url = f"{Config.BASE_URL}/ycdc/bakCmisYcOrgan/getCurrentIntegrityDetails?cecId={cec_id}"
    last_error = None

    for attempt in range(1, max_retries + 1):
        try:
            response = safe_request(session, detail_url)
            response_data = response.json()

            if response_data.get("code") != 0:
                print(f"信誉分明细接口异常: {response_data}")
                last_error = f"接口异常: {response_data}"
                continue

            encrypted_data = response_data.get("data", "")
            if not encrypted_data:
                print("信誉分明细接口返回空数据")
                last_error = "接口返回空数据"
                continue

            decrypted_str = aes_decrypt_base64(encrypted_data)
            detail_data = json.loads(decrypted_str)

            company_detail = {
                "cioName": detail_data.get("data", {}).get("cioName", company_name),
                "jfsj": detail_data.get("data", {}).get("jfsj", ""),
                "eqtName": detail_data.get("data", {}).get("eqtName", ""),
                "blxwArray": detail_data.get("data", {}).get("blxwArray", []),
                "lhxwArray": detail_data.get("data", {}).get("lhxwArray", []),
                "cecId": detail_data.get("data", {}).get("cecId", cec_id),
                "cechId": detail_data.get("data", {}).get("cechId", "")
            }
            print(f"成功获取企业信誉分明细: {company_detail.get('cioName')}")
            print(company_detail)
            return company_detail

        except Exception as e:
            print(f"第{attempt}次获取企业信誉分明细失败: {str(e)}")
            last_error = str(e)
            time.sleep(random.uniform(10, 30))

    print(f"获取企业信誉分明细失败: {last_error}")
    return {}


def fetch_company_details_with_cache(session: requests.Session, cec_id: str, company_name: str, cache: dict) -> dict:
    """带缓存的获取企业信誉分明细"""
    if cec_id in cache:
        print(f"使用缓存获取企业信誉分明细: {company_name}")
        return cache[cec_id]
    else:
        detail = fetch_company_detail(session, cec_id, company_name, max_retries=3)
        if detail:
            cache[cec_id] = detail
        return detail


def append_top_json(sorted_data: List[dict], category_name: str, github_mode: bool = False) -> Optional[str]:
    """追加数据到当天的JSON文件"""
    utc8_offset = timezone(timedelta(hours=8))
    now = datetime.now(utc8_offset)
    date_str = now.strftime("%Y%m%d")
    timestamp = now.strftime("%Y%m%d_%H%M%S")

    output_dir = os.getcwd()
    if github_mode:
        output_dir = os.path.join(output_dir, "excel_output")
        os.makedirs(output_dir, exist_ok=True)

    json_filename = f"{category_name}_top10.json"
    json_path = os.path.join(output_dir, json_filename)

    data_list = []
    for idx, item in enumerate(sorted_data[:10], 1):
        company_data = {
            "排名": idx,
            "企业名称": item.get("cioName", ""),
            "诚信分值": item.get("score", 0),
            "组织ID": item.get("orgId", ""),
        }
        if "detail" in item:
            company_data["信誉分明细"] = item["detail"]
        data_list.append(company_data)

    update_data = {
        "TIMEamp": timestamp,
        "DATAlist": data_list
    }

    if os.path.exists(json_path):
        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                existing_data = json.load(f)
            if not isinstance(existing_data, list):
                existing_data = [existing_data]
            existing_data.append(update_data)
        except:
            existing_data = [update_data]
    else:
        existing_data = [update_data]

    try:
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(existing_data, f, ensure_ascii=False, indent=2)
        print(f"已追加数据到JSON文件: {os.path.abspath(json_path)}")
        return json_path
    except Exception as e:
        print(f"JSON文件追加失败: {str(e)}")
        return None


def export_to_excel(data: List[dict], session: requests.Session, github_mode: bool = False) -> dict:
    """专业级Excel导出函数（多工作表分类排序）"""
    # -------------------- 数据处理 --------------------
    def process_item(item: dict) -> List[dict]:
        """将原始数据展开为明细行"""
        if item.get('eqtName') != '施工':
            return []

        main_info = {
            'cioName': item.get('cioName', ''),
            'eqtName': item.get('eqtName', ''),
            'csf': float(item.get('csf', 0)),
            'orgId': item.get('orgId', ''),
            'cecId': item.get('cecId', ''),
            'zzmx': ''
        }

        details = item.get('zzmxcxfArray', [])
        if not details:
            return [main_info]

        processed = []
        for detail in details:
            processed.append({
                **main_info,
                'zzmx': detail.get('zzmx', ''),
                'cxdj': detail.get('cxdj', ''),
                'score': float(detail.get('score', 0)),
                'jcf': float(detail.get('jcf', 0)),
                'zxjf': float(detail.get('zxjf', 0)),
                'kf': float(detail.get('kf', 0)),
                'eqlId': detail.get('eqlId', '')
            })
        return processed

    processed_data = []
    for item in data:
        if isinstance(item, dict):
            processed_data.extend(process_item(item))

    # -------------------- 创建工作簿 --------------------
    wb = Workbook()
    utc8_offset = timezone(timedelta(hours=8))
    timestamp = datetime.now(utc8_offset).strftime("%Y%m%d_%H%M%S")

    # 创建所有工作表
    summary_sheet = wb.active
    summary_sheet.title = Config.SHEET_CONFIGS[0]["name"]
    for config in Config.SHEET_CONFIGS[1:]:
        wb.create_sheet(title=config["name"])

    # -------------------- 样式配置 --------------------
    header_style = {
        'font': Font(bold=True, color="FFFFFF"),
        'fill': PatternFill("solid", fgColor="003366"),
        'alignment': Alignment(horizontal="center", vertical="center"),
        'border': Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
    }
    cell_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # -------------------- 填充数据 --------------------
    output_dir = os.getcwd()
    if github_mode:
        output_dir = os.path.join(output_dir, "excel_output")
        os.makedirs(output_dir, exist_ok=True)

    json_files = []
    detail_cache = {}

    for config in Config.SHEET_CONFIGS:
        ws = wb[config["name"]]
        ws.freeze_panes = config["freeze"]

        # 写入表头
        headers = [col['name'] for col in Config.COLUMNS]
        ws.append(headers)
        for col_idx, col in enumerate(Config.COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx)
            for attr, value in header_style.items():
                setattr(cell, attr, value)
            ws.column_dimensions[get_column_letter(col_idx)].width = col['width']

        # 准备数据
        if config["name"] == "企业信用数据汇总":
            sheet_data = processed_data
            merge_map = {}
        else:
            sheet_data = sorted(
                [d for d in processed_data
                 if str(d.get('zzmx', '')).startswith(config["prefix"])
                 and '级' in str(d.get('zzmx', ''))],
                key=lambda x: x.get('score', 0),
                reverse=True
            )
            print(f"过滤到数据量: {len(sheet_data)}")

            # 为前10名获取明细
            for item in sheet_data[:10]:
                cec_id = item.get('cecId')
                company_name = item.get('cioName')
                if not cec_id:
                    print(f"警告: 企业 {company_name} 缺少cecId，跳过")
                    continue
                detail = fetch_company_details_with_cache(session, cec_id, company_name, detail_cache)
                if detail:
                    item['detail'] = detail
                else:
                    print(f"警告: 未获取到企业 {company_name} 的信誉分明细")

            if config.get("generate_json"):
                print(f"\n正在生成 {config['name']} 的JSON排行榜...")
                json_path = append_top_json(sheet_data, config["name"], github_mode)
                if json_path:
                    json_files.append(json_path)

        if not sheet_data:
            print(f"警告: {config['name']} 无数据，跳过写入")
            continue

        # 写入数据行
        current_key = None
        start_row = 2
        merge_map = {} if config["merge"] else None

        for row_idx, row_data in enumerate(sheet_data, 2):
            if row_idx <= 4:
                print(f"写入行 {row_idx} 数据: {row_data['zzmx'][:20]}...")

            if config["merge"]:
                unique_key = f"{row_data['orgId']}-{row_data['cecId']}"
                if unique_key != current_key:
                    if current_key is not None:
                        merge_map[current_key] = (start_row, row_idx - 1)
                    current_key = unique_key
                    start_row = row_idx

            row = [row_data.get(col['id'], '') for col in Config.COLUMNS]
            ws.append(row)

            for col_idx in range(1, len(Config.COLUMNS) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = cell_border
                col_def = Config.COLUMNS[col_idx - 1]
                cell.alignment = Alignment(
                    horizontal=col_def['align'],
                    vertical='center',
                    wrap_text=False
                )
                if col_def.get('format'):
                    cell.number_format = col_def['format']

        # 合并单元格（汇总表）
        if config["merge"] and merge_map:
            if current_key:
                end_row = len(sheet_data) + 1
                if start_row <= end_row:
                    merge_map[current_key] = (start_row, end_row)
            for col in Config.COLUMNS:
                if col['merge']:
                    col_letter = get_column_letter(Config.COLUMNS.index(col) + 1)
                    for start, end in merge_map.values():
                        if end > start:
                            ws.merge_cells(f"{col_letter}{start}:{col_letter}{end}")

    # 删除默认空白工作表
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # 保存主文件
    filename = f"宜昌市信用评价信息_{timestamp}.xlsx" if github_mode else "宜昌市信用评价信息.xlsx"
    if github_mode:
        filename = os.path.join(output_dir, filename)
    try:
        wb.save(filename)
        print(f"文件已保存至：{os.path.abspath(filename)}")
        print("包含的工作表:")
        for sheet in wb.sheetnames:
            print(f"- {sheet}")
    except Exception as e:
        print(f"文件保存失败：{str(e)}")
        import traceback
        traceback.print_exc()
        return {"excel": None, "json": []}

    # -------------------- 生成信誉分明细表 --------------------
    print("\n=== 开始生成信誉分明细表（按资质类型精确匹配分值≥110） ===")
    try:
        cec_to_exact_qual_scores = {}
        cec_to_name = {}
        for record in processed_data:
            cec_id = record.get('cecId')
            if not cec_id:
                continue
            score = record.get('score', 0)
            if score < 110:
                continue
            qual_name = record.get('zzmx', '')
            if not qual_name:
                continue
            company_name = record.get('cioName', '')
            if cec_id not in cec_to_exact_qual_scores:
                cec_to_exact_qual_scores[cec_id] = {}
                cec_to_name[cec_id] = company_name
            if qual_name not in cec_to_exact_qual_scores[cec_id] or score > cec_to_exact_qual_scores[cec_id][qual_name]:
                cec_to_exact_qual_scores[cec_id][qual_name] = score

        if not cec_to_exact_qual_scores:
            print("没有诚信分值≥110的企业，跳过信誉分明细表生成。")
        else:
            # 获取明细
            for cec_id in cec_to_exact_qual_scores.keys():
                if cec_id not in detail_cache:
                    time.sleep(random.uniform(5, 15))
                    company_name = cec_to_name.get(cec_id, '')
                    detail = fetch_company_detail(session, cec_id, company_name, max_retries=3)
                    if detail:
                        detail_cache[cec_id] = detail
                    else:
                        print(f"警告: 无法获取企业 {company_name} 的信誉分明细，跳过该企业。")
                        detail_cache[cec_id] = None

            # 创建明细工作簿
            detail_wb = Workbook()
            default_sheet = detail_wb.active
            detail_wb.remove(default_sheet)

            bad_headers = [
                "企业名称", "诚信分值", "违规人员", "身份证号", "违规事由", "项目名称",
                "资质类型", "行为类别", "开始日期", "结束日期", "有效期 (月)", "扣分值", "确认书编号"
            ]
            good_headers = [
                "企业名称", "诚信分值", "获奖 / 表彰事由", "项目名称",
                "资质类型", "行为类别", "开始日期", "结束日期", "有效期 (月)", "加分值", "文号"
            ]

            bad_sheet = detail_wb.create_sheet("不良行为")
            bad_sheet.append(bad_headers)
            good_sheet = detail_wb.create_sheet("良好行为")
            good_sheet.append(good_headers)

            # 填充数据
            for cec_id, qual_scores in cec_to_exact_qual_scores.items():
                detail = detail_cache.get(cec_id)
                if not detail:
                    continue
                company_name = cec_to_name.get(cec_id, '')

                for bl in detail.get('blxwArray', []):
                    qual_type = bl.get('kfqyzz', '')
                    if not qual_type:
                        continue
                    matched_score = qual_scores.get(qual_type)
                    if matched_score is None:
                        print(f"警告: 企业 {company_name} 不良行为关联资质 '{qual_type}' 未精确匹配到分值≥110的资质，跳过该行为。")
                        continue
                    bad_sheet.append([
                        company_name, matched_score,
                        bl.get('cfry', ''), bl.get('cfryCertNum', ''), bl.get('reason', ''),
                        bl.get('engName', ''), qual_type, bl.get('bzXwlb', ''),
                        bl.get('beginDate', ''), bl.get('endDate', ''), bl.get('valid', ''),
                        bl.get('realValue', 0), bl.get('kftzsbh', '')
                    ])

                for lh in detail.get('lhxwArray', []):
                    qual_type = lh.get('jfqyzz', '')
                    if not qual_type:
                        continue
                    matched_score = qual_scores.get(qual_type)
                    if matched_score is None:
                        print(f"警告: 企业 {company_name} 良好行为关联资质 '{qual_type}' 未精确匹配到分值≥110的资质，跳过该行为。")
                        continue
                    proj_name = lh.get('engName', '') or lh.get('hjyy', '')
                    good_sheet.append([
                        company_name, matched_score,
                        lh.get('reason', ''), proj_name, qual_type,
                        lh.get('bzXwlb', ''), lh.get('beginDate', ''), lh.get('endDate', ''),
                        lh.get('valid', ''), lh.get('realValue', 0), lh.get('documentNumber', '')
                    ])

            # 应用样式
            header_fill = PatternFill("solid", fgColor="003366")
            header_font = Font(bold=True, color="FFFFFF")
            header_alignment = Alignment(horizontal="center", vertical="center")
            header_border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin")
            )
            data_alignment = Alignment(horizontal="center", vertical="center")
            data_border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin")
            )

            for sheet in [bad_sheet, good_sheet]:
                sheet.freeze_panes = 'A2'
                for cell in sheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.border = header_border
                for row in sheet.iter_rows(min_row=2):
                    for cell in row:
                        cell.alignment = data_alignment
                        cell.border = data_border
                for col in sheet.columns:
                    max_len = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        if cell.value:
                            content = str(cell.value)
                            length = sum(2 if '\u4e00' <= c <= '\u9fff' else 1 for c in content)
                            if length > max_len:
                                max_len = length
                    adjusted_width = min(max(max_len + 2, 8), 50)
                    sheet.column_dimensions[col_letter].width = adjusted_width

            # 保存明细表
            detail_filename = f"信誉分明细表_{timestamp}.xlsx"
            if github_mode:
                detail_filename = os.path.join(output_dir, detail_filename)
            else:
                detail_filename = os.path.join(os.getcwd(), detail_filename)
            detail_wb.save(detail_filename)
            print(f"信誉分明细表已保存至：{os.path.abspath(detail_filename)}")
            print(f"不良行为记录数：{bad_sheet.max_row-1}")
            print(f"良好行为记录数：{good_sheet.max_row-1}")
    except Exception as e:
        print(f"生成信誉分明细表时发生错误: {str(e)}")
        import traceback
        traceback.print_exc()

    return {"excel": filename, "json": json_files}


def main():
    print("=== 启动数据获取程序 ===")
    session = requests.Session()
    all_data = []

    try:
        current_code, current_ts = get_new_code(session)
        print(f"[初始化] 验证码: {current_code} | 时间戳: {current_ts}")

        first_data, total = process_page(session, 1, current_code, current_ts)
        total_pages = (total + Config.PAGE_SIZE - 1) // Config.PAGE_SIZE
        print(f"[初始化] 总记录数: {total} | 总页数: {total_pages}")

        if total == 0:
            print("错误: API返回总记录数为0，无需继续处理")
            return

        page = 1
        while page <= total_pages:
            retry_count = 0
            success = False

            while retry_count < Config.PAGE_RETRY_MAX and not success:
                try:
                    print(f"\n[处理中] 第 {page} 页 (重试次数: {retry_count})")
                    page_data, _ = process_page(session, page, current_code, current_ts)
                    if page_data:
                        print(f"[成功获取数据] 第 {page} 页 {len(page_data)} 条记录")
                        all_data.extend(page_data)
                        success = True
                        page += 1
                    else:
                        print(f"[警告] 第 {page} 页获取到空数据，尝试刷新验证码")
                        raise RuntimeError("empty page data")
                except Exception as e:
                    retry_count += 1
                    print(f"[重试] 第 {page} 页第 {retry_count} 次重试: {str(e)}")
                    try:
                        current_code, current_ts = get_new_code(session)
                        print(f"[刷新] 新验证码: {current_code} | 新时间戳: {current_ts}")
                    except Exception as e:
                        print(f"[警告] 验证码刷新失败: {str(e)}")
                        break

            if not success:
                print(f"[终止] 第 {page} 页超过最大重试次数，跳过此页")
                page += 1

        print(f"\n=== 数据获取完成 ===")
        print(f"总获取记录数: {len(all_data)}")

        if all_data:
            export_result = export_to_excel(all_data, session, github_mode=True)
            if export_result:
                json_files = export_result.get("json", [])
                github_output = os.getenv('GITHUB_OUTPUT')
                if github_output:
                    with open(github_output, 'a') as f:
                        f.write(f'excel-path={export_result["excel"]}\n')
                    for i, json_path in enumerate(json_files, 1):
                        with open(github_output, 'a') as f:
                            f.write(f'json-path-{i}={json_path}\n')
                else:
                    print("::注意:: 未在GitHub Actions环境中，跳过输出设置")

                print("\n=== 所有生成的文件 ===")
                print(f"Excel文件: {export_result['excel']}")
                for i, json_path in enumerate(json_files, 1):
                    print(f"JSON文件 #{i}: {json_path}")
        else:
            print("错误: 没有获取到任何有效数据，无法导出Excel")
    except Exception as e:
        print(f"\n!!! 程序执行失败 !!!\n错误原因: {str(e)}")
        import traceback
        traceback.print_exc()
    finally:
        session.close()


if __name__ == "__main__":
    main()
